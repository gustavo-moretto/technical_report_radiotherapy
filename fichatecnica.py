import os


def fichatecnica(arq, font_size, row_height,deletar = False):
    import openpyxl
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import datetime
    import requests
    import json
    import os

    # selecionar arquivo desejado e criar lista para colocar as informações
    arq = (arq)
    rtpfile = open(f'\\\\Srvmosaiqapp\\mosaiq_app\\RTP\\{arq}', 'r')
    dados = []
    fichatec = []
    for i in rtpfile:
        dados.append(i)

    # retirar espaços sem informação no arquivo gerado
    # porém, eles não são deletados, pois são úteis para verificar a aprovação do plano
    j = 0
    while j < len(dados):
        dados[j] = dados[j].split(',')
        z = 0
        while z < len(dados[j]):
            dados[j][z] = dados[j][z].strip('""')
            z += 1
        j += 1

    e = 0
    # mostra o nome e o ID do paciente para que seja elaborada a ficha técnica
    rtpfile.close()
    ID = dados[0][1]
    Nome = dados[0][2]
    plano_aprov = dados[0][5]
    print(f'Paciente: {Nome}; ID: {ID}; Plano aprovado: {plano_aprov}')
    # imprime dados de aprovação do plano
    if dados[0][13] != '':
        Revisor = dados[0][13]
        Data_aprov = f'{dados[0][6][-2:]}/{dados[0][6][-4:-2:1]}/{dados[0][6][-8:-4:1]}'
        ID = dados[0][1]
        Nome = dados[0][2]
        plano_aprov = dados[0][5]
        # aqui ele obterá o diagnóstico baseado no CID10
        x = ''
        try:
            tabela = requests.get("http://cid.api.mokasoft.org/cid10/")
            CID10 = tabela.json()
            cod = str(input('Digite o CID10: ')).upper()
            b = int(0)
            if len(cod) == 4:
                cod = cod[:]
                cod = list(cod)
                cod = cod[0] + cod[1] + cod[2] + '.' + cod[3]
            while b < len(CID10):
                if cod in CID10[b]['codigo']:
                    y = CID10[b].get('codigo')
                    z = CID10[b].get('nome').title()
                    x = 'CID encontrado'
                    break
                else:
                    b += 1
            Diagnóstico = y + ' - ' + z
        except:
            # caso ele não encontre o CID10 na API, deverá ser digitado manualmente
            if x == '':
                print('CID não encontrado ou não é possível acessar a API')
                Diagnóstico = str(input('Digite o CID e o diagnóstico manualmente:'))
        finally:
            Finalidade = str(input('Digite a finalidade da radioterapia: '))
            fichatec.append(f'Aprovado por {Revisor}, em {Data_aprov}.')
            fichatec.append(f'Plano aprovado no Monaco: {plano_aprov}')
            fichatec.append(f'Paciente: {Nome}; ID: {ID}')
            fichatec.append(f'CID10 e Diagnóstico: {Diagnóstico.title()}; Finalidade: {Finalidade.title()}')
            fichatec.append(f'Quimioterapia concomitante: SIM (  )       NÃO (   )')
            data = datetime.datetime.now()
            data_texto = data.strftime('%d/%m/%Y; %H:%M')
            fichatec.append(f'Ficha técnica gerada em: {data_texto}.')
            fichatec.append('---' * 80)

        n = 0
        posic = ''
        for t in dados:
            if n == 0:
                if 'SITE_SETUP_DEF' in t:
                    n += 1
                    AL = t[3]
                    if t[2] == 'HFS':
                        posic = 'Decúbito dorsal; cabeça para o gantry'
                    elif t[2] == 'HFP':
                        posic = 'Decúbito ventral; cabeça para o gantry'
                    elif t[2] == 'FFS':
                        posic = 'Decúbito dorsal; pés para o gantry'
                    else:
                        posic = 'Decúbito ventral; pés para o gantry'
                    fichatec.append(f'POSICIONAMENTO: {posic}')
                    acessorios = str(input(f'Insira os acessórios utilizados: '))
                    fichatec.append(f'ACESSÓRIOS: {acessorios}')
                    fichatec.append(f'Acelerador linear: {AL}')
                    # fichatec.append('---' * 70)
                    info = ''
                    add_info = str(input('Deseja adicionar alguma informação [S/N]: '))
                    if add_info == 'S' or add_info == 's':
                        info = str(input('Digite a informação a ser adicionada: '))
                        fichatec.append(f'Informações adicionais: {info}')
                        # fichatec.append('---' * 70)
                    else:
                        fichatec.append('Informações adicionais: ')
                        # fichatec.append('---' * 70)

        n = 0
        t = 0
        fracionamento = ''
        # armazena dados da(s) prescrição(ões) na lista 'presc'
        for t in dados:
            if 'RX_DEF' in t:
                fichatec.append('---' * 80)
                presc = f'{t[2][:]}'
                fracoes = float(t[7]) / float(t[8])
                change_presc_name = str(input(f'O nome da prescrição é {presc}; {t[7]}cGy/{round(fracoes)}fr. Caso deseje alterá-lo, digite A: '))
                if change_presc_name == 'A' or change_presc_name == 'a':
                    presc = str(input('Digite o nome que deseja para esta prescrição: '))
                    dose_levels = str(input('Há mais de um nível de dose [S/N]? '))
                    dados_aprov = (f'{presc}: {t[7]}cGy/{round(fracoes)} frações; {t[8]}cGy/fração.')
                    if dose_levels == 's' or dose_levels == 'S':
                        smart = str(input('Digite os níveis de dose: '))
                        dados_aprov = (
                            f'{presc}: {t[7]}cGy/{round(fracoes)} frações; {t[8]}cGy/fração. Níveis de dose: {smart}.')
                    schedule = str(input('O esquema de fracionamento é diferente de diário [S/N]? '))
                    fracionamento = ''
                    if schedule == 'S' or schedule == 's':
                        fracionamento = str(input(f'Digite o esquema de fracionamento ({presc}):'))
                else:
                    dados_aprov = (f'{presc}: {t[7]}cGy/{round(fracoes)} frações; {t[8]}cGy/fração.')
                    dose_levels = str(input('Há mais de um nível de dose [S/N]? '))
                    if dose_levels == 's' or dose_levels == 'S':
                        smart = str(input('Digite os níveis de dose: '))
                        dados_aprov = (
                            f'{presc}: {t[7]}cGy/{round(fracoes)} frações; {t[8]}cGy/fração. Níveis de dose: {smart}.')
                    schedule = str(input('O esquema de fracionamento é diferente de diário [S/N]? '))
                    fracionamento = ''
                    if schedule == 'S' or schedule == 's':
                        fracionamento = str(input(f'Digite o esquema de fracionamento ({presc}):'))
                fichatec.append(f'{dados_aprov}')
                if fracionamento != '':
                    fichatec.append(f'Esquema de fracionamento: {fracionamento}')
                # fichatec.append('---' * 67)
            if 'SITE_SETUP_DEF' in t:
                fichatec.append('---' * 80)
                Campos = 'CAMPO'
                UMt = 'UMtotal'
                UMf = 'UMfiltro'
                G = 'Gantry/Arco'
                C = 'Colimador'
                M = 'Mesa'
                modo = 'Modalidade'
                Energia = 'Energia'
                dist = 'SSD(cm)'
                field_size = '(X/Y)'
                cone_size = 'Cone'
                bolus = 'Bólus'
                cabecalho = f'{Campos} {Energia} {UMf}/{UMt} {modo} {dist} {cone_size} {bolus} {G} {C} {M} {field_size}'
                fichatec.append(cabecalho.split())
            if 'FIELD_DEF' in t:
                # para não ocorrer erros na hora de juntar os campos, criei o 'field_data' que une o 'field_number' e o 'field_name'
                # e substitui os espaços, deixando sem espaço entre eles, para não separar na hora dar o comando 'split()'
                field_name = t[2]
                if field_name == '':
                    field_name = input(f'Digite o nome do campo (G: {t[16]}º;C: {t[17]}º;M: {t[29]}º: ')
                field_number = t[3]
                field_data = field_number + '-' + field_name
                field_data = field_data.replace(" ", "")
                if t[9] == 'Static' and t[10] == 'Xrays':
                    n += 1
                    field_UM = float(t[6])
                    field_UM = f'{field_UM:.2f}'
                    cone = 'NA'
                    field_bolus = str(input(f'Campo {field_name} possui bólus [S/N]: '))
                    if field_bolus == 's' or field_bolus == 'S':
                        field_bolus = str(input('Digite a espessura do bólus: '))
                    else:
                        field_bolus = 'NA'
                    if t[7] != "":
                        field_UMfiltro = t[7]
                    else:
                        field_UMfiltro = 0.0
                    field_energy = f'{t[11]}MV'
                    field_SSD = f'{t[15]}'
                    field_modality = '3D'
                    field_gantry = f'{t[16]}'
                    field_col = f'{t[17]}º'
                    field_couch = f'{t[29]}º'
                    if t[20] != '' and t[21] != '':
                        field_X = f'{abs(float(t[20])) + float(t[21])}'
                        field_X = f'{float(field_X):.1f}'
                        field_Y = f'{abs(float(t[24])) + float(t[25])}'
                        field_Y = f'{float(field_Y):.1f}'
                    else:
                        field_X = '40.0'
                        field_Y = f'{abs(float(t[24])) + float(t[25])}'
                        field_Y = f'{float(field_Y):.1f}'
                elif t[9] == 'Static' and t[10] == 'Elect':
                    n += 1
                    field_UM = float(t[6])
                    field_UM = f'{field_UM:.2f}'
                    field_bolus = str(input(f'Campo {field_name} possui bólus: '))
                    if field_bolus == 's' or field_bolus == 'S':
                        field_bolus = str(input('Digite a espessura do bólus: '))
                    else:
                        field_bolus = 'NA'
                    if t[7] != "":
                        field_UMfiltro = t[7]
                    else:
                        field_UMfiltro = 0.0
                    field_energy = f'{t[11]}MeV'
                    field_SSD = f'{t[15]}'
                    field_modality = 'Elétrons'
                    field_gantry = f'{t[16]}'
                    field_col = f'{t[17]}º'
                    field_couch = f'{t[29]}º'
                    field_X = '-'
                    field_Y = '-'
                    cone = str(t[40])
                elif t[9] == 'Setup':
                    field_bolus = 'NA'
                    field_modality = 'Setup'
                    field_SSD = f'{t[15]}'
                    field_gantry = f'{t[16]}'
                    field_col = f'{t[17]}º'
                    field_couch = f'{t[29]}º'
                    field_UM = '-'
                    field_UMfiltro = '-'
                    field_X = '-'
                    field_Y = '-'
                    cone = '-'
                # o else fica para campos 'Dynamic'
                else:
                    n += 1
                    field_UM = float(t[6])
                    field_UM = f'{field_UM:.2f}'
                    field_bolus = str(input(f'Campo {field_name} possui bólus [S/N]: '))
                    if field_bolus == 's' or field_bolus == 'S':
                        field_bolus = str(input('Digite a espessura do bólus: '))
                    else:
                        field_bolus = 'NA'
                    if t[7] != "":
                        field_UMfiltro = t[7]
                    else:
                        field_UMfiltro = 0.0
                    field_energy = f'{t[11]}MV'
                    field_SSD = f'{t[15]}'
                    vmat_arcdyn = str(input('Digite V para VMAT ou D para arco dinâmico: '))
                    while True:
                        if vmat_arcdyn == 'v' or vmat_arcdyn == 'V':
                            field_modality = 'VMAT'
                            break
                        elif vmat_arcdyn == 'd' or vmat_arcdyn == 'D':
                            field_modality = 'ArcoDinâmico'
                            break
                        else:
                            print('Entrada inválida!')
                            vmat_arcdyn = str(input('Digite V para VMAT ou D para arco dinâmico: '))
                    field_gantry = f'{t[16]}'
                    field_col = f'{t[17]}º'
                    field_couch = f'{t[29]}º'
                    field_rotation = f'{t[32]}'
                    gantry_start = f'{t[33]}'
                    gantry_final = f'{t[34]}'
                    cone = 'NA'
                    field_X = '-'
                    field_Y = '-'
                    #extrai os dados de rotação para vmat ou arcdyn
                    angle_rotation = ''
                    start_angle = ''
                    final_angle = ''
                    e += 1
                    for p in dados:
                        if "CONTROL_PT_DEF" in p:
                            if p[1] == field_number or p[1] == e:
                                if p[5] == '0':
                                    start_angle = f'{float(p[13]):.2f}'
                                elif p[14] == '' or p[14] != field_rotation and int(p[5]) >= float((int(p[4])/2)):
                                    final_angle = f'{float(p[13]):.2f}'
                                    break
                    # no if, define o ângulo de rotação do campo
                    if gantry_start == gantry_final:
                        if field_rotation == 'CW':
                            if 180 <= float(start_angle) < 360 and 0 <= float(final_angle) <= 180:
                                angle_rotation = float(final_angle) + (360 - float(start_angle))
                            else:
                                angle_rotation = float(final_angle) - float(start_angle)
                        else:
                            if 0 <= float(start_angle) <= 180 and 180 <= float(final_angle) < 360:
                                angle_rotation = -float(start_angle) - (360 - float(final_angle))
                            else:
                                angle_rotation = float(final_angle) - float(start_angle)
                    else:
                        if field_rotation == 'CW':
                            if 180 <= float(gantry_start) < 360 and 0 <= float(gantry_final) <= 180:
                                angle_rotation = (360 - float(gantry_start)) + float(gantry_final)
                            else:
                                angle_rotation = float(gantry_final) - float(gantry_start)
                        else:
                            if 0 <= float(gantry_start) <= 180 and 180 <= float(gantry_final) < 360:
                                angle_rotation = -float(gantry_start) - (360 - float(gantry_final))
                            else:
                                angle_rotation = float(gantry_final) - float(gantry_start)
                if field_modality == '3D':
                    dados_campos = f'({n}){field_data:5} {field_energy:12} {field_UMfiltro:.6}/{field_UM:10} {field_modality:9}  {field_SSD:10} {cone} {field_bolus} {field_gantry}º {field_col} {field_couch:12} {field_X}/{field_Y}'
                    fichatec.append(dados_campos.split())
                elif field_modality == 'Elétrons':
                        dados_campos = f'({n}){field_data:5} {field_energy:12} {field_UMfiltro:.6}/{field_UM:10} {field_modality:9}  {field_SSD:10} {cone} {field_bolus} {field_gantry}º {field_col} {field_couch:12} {field_X}/{field_Y}'
                        fichatec.append(dados_campos.split())
                elif field_modality == 'VMAT' or field_modality == 'ArcoDinâmico':
                    dados_campos = f'({n}){field_data:5} {field_energy:12} {field_UMfiltro:.5}/{field_UM:10} {field_modality:9} {field_SSD:10} {cone} {field_bolus} {gantry_start}º/{angle_rotation}º {field_col} {field_couch:12} {field_X}/{field_Y}'
                    fichatec.append(dados_campos.split())
                else:
                    dados_campos = f'{field_data:5} {field_energy:12} {field_UMfiltro:.5}/{field_UM:10} {field_modality:9}  {field_SSD:10} {cone} {field_bolus} {field_gantry}º {field_col} {field_couch:12} {field_X}/{field_Y}'
                    fichatec.append(dados_campos.split())

        # abre o arquivo 'fichatecnica' e seleciona a Sheet 'Dados ficha tec'
        l = 0
        while l < 2:
            fichatec.append('---' * 80)
            fichatec.append(' ')
            fichatec.append('---' * 80)
            fichatec.append('---' * 80)
            l += 1

        wb = openpyxl.load_workbook("fichatecnica.xlsx")
        sh = wb['Check-list']
        img = openpyxl.drawing.image.Image('logo.tif')
        img.anchor = 'A1'
        sh.add_image(img)
        sh = wb['Dados ficha tec']
        img = openpyxl.drawing.image.Image('logo.tif')
        img.anchor = 'A1'
        sh.add_image(img)
        y = 6
        # imprime linha a linha os dados da ficha técnica
        # caso a linha seja o cabeçalho das informações do campo ou as informações do campo,
        # entra em um laço de repetição, para alternar a célula a cada impressão
        for t in fichatec:
            if t == cabecalho.split():
                l = 1
                for k in t:
                    sh.cell(row=y, column=l, value=k)
                    cell = sh.cell(row=y, column=l, value=k)
                    cell.font = Font(size=font_size)
                    cell.border = Border(bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='center')
                    sh.row_dimensions[y].height = row_height
                    l += 1
                y += 1
            elif len(t[0]) > 1:
                l = 1
                for k in t:
                    sh.cell(row=y, column=l, value=k)
                    cell = sh.cell(row=y, column=l, value=k)
                    cell.font = Font(size=font_size)
                    cell.alignment = Alignment(horizontal='center')
                    sh.row_dimensions[y].height = row_height
                    l += 1
                y += 1
            else:
                sh.cell(row=y, column=1, value=t)
                cell = sh.cell(row=y, column=1, value=t)
                cell.font = Font(bold=True)
                y += 1

        # a partir daqui, ocorre a configuração do arquivo excel da ficha técnica
        sh.column_dimensions['A'].width = 21
        sh.column_dimensions['B'].width = 12
        sh.column_dimensions['C'].width = 16
        sh.column_dimensions['D'].width = 13
        sh.column_dimensions['E'].width = 11
        sh.column_dimensions['F'].width = 13
        sh.column_dimensions['G'].width = 13
        sh.column_dimensions['H'].width = 13
        sh.column_dimensions['I'].width = 16
        cell = sh.cell(row=9, column=1)
        cell.font = Font(italic=True, bold=True)
        sh['A9'].font = Font(size=10, bold=True)
        sh.merge_cells('A6:D6')
        sh.merge_cells('A7:D7')
        sh.merge_cells('A8:D8')
        sh.merge_cells('A10:D10')
        sh.merge_cells('A12:J12')
        sh.merge_cells('A13:J13')
        sh.merge_cells('A14:J14')

        # extrair dados para calculo paralelo
        calc = []
        k = 0
        field_X = ''
        field_Y = ''
        fracoes = 0
        for k in dados:
            field_dose = ''
            if 'RX_DEF' in k:
                fracoes = float(k[7]) / float(k[8])
            if 'FIELD_DEF' in k:
                # para não ocorrer erros na hora de juntar os campos, criei o 'field_data' que une o 'field_number' e o 'field_name'
                # e substitui os espaços, deixando sem espaço entre eles, para não separar na hora dar o comando 'split()'
                field_name = k[2]
                field_number = k[3]
                field_data = field_number + '-' + field_name
                field_data = field_data.replace(" ", "")
                field_dose = float(k[5])
                field_UM = f'{(k[6]):.5}'
                if k[7] != "":
                    field_UMfiltro = k[7]
                else:
                    field_UMfiltro = 0.0
                field_energy = f'{k[11]}'
                field_SSD = f'{k[15]}'
                if t[9] == 'Static' and 'Xrays':
                    if k[20] != '' and k[21] != '':
                        field_X = f'{abs(float(k[20])) + float(k[21])}'
                        field_X = f'{float(field_X):.1f}'
                        field_Y = f'{abs(float(k[24])) + float(k[25])}'
                        field_Y = f'{float(field_Y):.1f}'
                    else:
                        field_X = '40.0'
                        field_Y = f'{abs(float(k[24])) + float(k[25])}'
                        field_Y = f'{float(field_Y):.1f}'
                dados_campos = f'{field_data:5} {field_energy:12} ={field_dose}*{fracoes} {field_UMfiltro:.5}/{field_UM:10} {field_SSD:10} {field_X}/{field_Y}'
                # dados_campos = dados_campos.replace(".", ",")
                calc.append(dados_campos.split())

        sh = wb['Calculo paralelo']
        y = 1
        for t in calc:
            if len(t[0]) > 1:
                l = 1
                for k in t:
                    sh.cell(row=l, column=y, value=k)
                    cell = sh.cell(row=l, column=y, value=k)
                    cell.alignment = Alignment(horizontal='center')
                    l += 1
                y += 1

        # salva a planilha com o nome, ID e plano aprovado
        sh = wb['Dados ficha tec']
        ws = wb.active
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        # ws.page_setup.fitToHeight = True
        # ws.page_setup.fitToWidth = True
        wb.save(filename=f'C:\\Users\\Elekta\\Desktop\\Ficha tecnica\\Fichas tecnicas para impressão\\{Nome}-{ID}, {plano_aprov}.xlsx')
        wb.close()

        path = f"C:\\Users\\Elekta\\Desktop\\Ficha tecnica\\Fichas tecnicas para impressão\\{Nome}-{ID}, {plano_aprov}.xlsx"
        path = os.path.realpath(path)
        os.startfile(path)

        if '3D' in fichatec:
            path = f"C:\\Users\\Elekta\\Desktop\\CÁLCULO PARALELO"
            path = os.path.realpath(path)
            os.startfile(path)

        if deletar:
            os.remove(f'\\\\Srvmosaiqapp\\mosaiq_app\\RTP\\{arq}')


    else:
        print('Plano não aprovado! \nNão é possível gerar a ficha técnica.')
